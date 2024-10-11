import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import Select
from openpyxl import load_workbook
import time

# WordPress login credentials
wp_url = 'https://aheart2help.com/wp-login.php?redirect_to=https%3A%2F%2Faheart2help.com%2Fwp-admin%2Fpost-new.php%3Fpost_type%3Dorganization&reauth=1'
wp_username = ''
wp_password = ''

# Path to the Excel file
excel_file_path = 'delaware_2.xlsx'

# Load the workbook and select the active sheet
wb = load_workbook(excel_file_path)
sheet = wb.active

# Initialize the WebDriver
driver = webdriver.Chrome(service=Service('chromedriver.exe'))

       
try:
    driver.get(wp_url)

    # Log in to WordPress
    username = driver.find_element(By.ID, 'user_login')
    password = driver.find_element(By.ID, 'user_pass')
    username.send_keys(wp_username)
    password.send_keys(wp_password)

    # Optionally, check the "Remember Me" checkbox
    remember_me = driver.find_element(By.ID, 'rememberme')
    if not remember_me.is_selected():
        remember_me.click()

    # Submit the form
    submit_button = driver.find_element(By.ID, 'wp-submit')
    submit_button.click()

    # Wait for the login to complete
    WebDriverWait(driver, 10).until(EC.url_changes(wp_url))

    # Navigate to the page with the form where you need to input the Excel data
    driver.get('https://aheart2help.com/wp-admin/post-new.php?post_type=organization')

    # Wait for the form elements to be present
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, 'acf-group_65e1dc59d2db8')))
    
        # Iterate through the rows in the Excel sheet and fill out the form
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
        try:
            data = list(row)

            # Fill out the title field (post title)
            driver.find_element(By.ID, 'title').send_keys(data[0])  # Title
            
            # Fill out the form fields
            driver.find_element(By.ID, 'acf-field_65e212550935c').send_keys(data[1])  # Organization Name
            driver.find_element(By.ID, 'acf-field_65e57914ad0d4').send_keys(data[2])  # In Care Of Name
            state_dropdown = Select(driver.find_element(By.ID, 'acf-field_65e67db7810ee'))
            state_dropdown.select_by_visible_text(data[3])  # Select state by visible text
            driver.find_element(By.ID, 'acf-field_65e1dc5b3129a').send_keys(data[4])  # City
            driver.find_element(By.ID, 'acf-field_65e20c5509914').send_keys(data[5])  # Zip Code
            driver.find_element(By.ID, 'acf-field_65e1dc8f3129b').send_keys(data[6])  # EIN
            driver.find_element(By.ID, 'acf-field_65e5701dfe6b9').send_keys(data[7])  # NTEE Common Code
            driver.find_element(By.ID, 'acf-field_65e5715beb15e').send_keys(data[8])  # Organization Code
            driver.find_element(By.ID, 'acf-field_65e57181eb15f').send_keys(data[9])  # Deductibility Code
            driver.find_element(By.ID, 'acf-field_65e571a44f0a0').send_keys(data[10])  # Affiliation Code
            driver.find_element(By.ID, 'acf-field_65e571bc4f0a1').send_keys(data[11])  # Subsection/classification Codes
            driver.find_element(By.ID, 'acf-field_65e571d64f0a2').send_keys(data[12])  # Activity Codes
            driver.find_element(By.ID, 'acf-field_65e571fc9e7bc').send_keys(data[13])  # NTEE Code
            driver.find_element(By.ID, 'acf-field_65e5721e9e7bd').send_keys(data[14])  # Foundation Code
            driver.find_element(By.ID, 'acf-field_65e5723f610f7').send_keys(data[15])  # Exempt Organization Status Code
            driver.find_element(By.ID, 'acf-field_65e1dca33129c').send_keys(data[16])  # Income/Assets
            driver.find_element(By.ID, 'acf-field_65e5725f610f8').send_keys(data[17])  # Tax Period
            driver.find_element(By.ID, 'acf-field_65e573255efb6').send_keys(data[18])  # Ruling Date
            driver.find_element(By.ID, 'acf-field_65e5735528416').send_keys(data[19])  # Accounting Period
            driver.find_element(By.ID, 'acf-field_65e573d3ec772').send_keys(data[20])  # Asset Code
            driver.find_element(By.ID, 'acf-field_65e575198b93f').send_keys(data[21])  # Income Code
            driver.find_element(By.ID, 'acf-field_65e5754177e46').send_keys(data[22])  # Asset Amount
            driver.find_element(By.ID, 'acf-field_65e5755e13167').send_keys(data[23])  # Income Amount
            driver.find_element(By.ID, 'acf-field_65e5760fff28c').send_keys(data[24])  # Filing Requirement Code
            driver.find_element(By.ID, 'acf-field_65e57636c81f7').send_keys(data[25])  # Form 990 Revenue Amount
            driver.find_element(By.ID, 'acf-field_65e5765b4651b').send_keys(data[26])  # PF Filing Requirement Code
            driver.find_element(By.ID, 'acf-field_65e57879455f1').send_keys(data[27])  # Google Map Link For Get Direction
            driver.find_element(By.ID, 'acf-field_65e576f5cf203').send_keys(data[28])  # Demographic Address
            driver.find_element(By.ID, 'acf-field_65e57a31623ab').send_keys(data[29])  # Complete Valid Organization Address In Text For Map Preview
            
            time.sleep(10)
 
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, 'org-category-all')))
            
            # Click on the "All Org Categories" tab if it's not active
            #all_org_categories_tab = driver.find_element(By.XPATH, "//a[@href='#org-category-all']")
            #if not 'current' in all_org_categories_tab.get_attribute("class"):
            # all_org_categories_tab.click()


            # Select the "State" category
            state_checkbox = WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'in-org-category-44')))
            if not state_checkbox.is_selected():
                state_checkbox.click()

            # Select the "Delaware" category
            delaware_checkbox = WebDriverWait(driver,10).until(EC.element_to_be_clickable((By.ID, 'in-org-category-56')))
            if not delaware_checkbox.is_selected():
                delaware_checkbox.click()

            # Optionally: Print status
            print("Selected 'State' and 'Alabama' categories.")


            # Save the draft
            save_draft_button = driver.find_element(By.ID, 'save-post')
            save_draft_button.click()
            
            time.sleep(10)
            
             # Publish the post
             # Publish the post
            publish_button = driver.find_element(By.ID, 'publish')
            publish_button.click()
        
        # Wait for the post to be published
            #WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, 'message')))
            time.sleep(5)  # Additional wait to ensure the post is published before proceeding
            
            driver.get('https://aheart2help.com/wp-admin/post-new.php?post_type=organization')

            
            # Wait for the form submission to process and for the page to reload
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, 'acf-group_65e1dc59d2db8')))
            time.sleep(5)  # Additional wait to ensure the draft is saved before the next iteration
            

        except Exception as e:
            print(f"Error processing row: {data}. Exception: {e}")

finally:
    # Close the browser
    driver.quit()
